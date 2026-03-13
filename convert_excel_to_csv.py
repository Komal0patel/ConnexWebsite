
import openpyxl
import csv
import os

def process_excel(file_path, output_csv):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    all_data = []
    
    # Headers we expect based on the screenshot/analysis
    headers = [
        'Category', 'SL No.', 'Product Name', 'Product Code', 'Type', 
        'Diameter', 'Height', 'Length', 'Width', 'Power (W)', 
        'Lumens (L/W)', 'Mounting', 'Light Distribution', 'Light Source', 
        'IP', 'CRI', 'CCT (K)', 'Comments'
    ]
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"Processing sheet: {sheet_name}")
        
        # Find the header row. It usually contains 'PRODUCT CODE' or 'SL No.'
        header_row_idx = -1
        for row in sheet.iter_rows(min_row=1, max_row=10):
            row_values = [cell.value for cell in row]
            if any(val and isinstance(val, str) and 'PRODUCT CODE' in val.upper() for val in row_values):
                header_row_idx = row[0].row
                break
        
        if header_row_idx == -1:
            print(f"Skipping sheet {sheet_name}: No header found")
            continue
            
        # Map columns
        col_map = {}
        for cell in sheet[header_row_idx]:
            val = str(cell.value).strip().upper() if cell.value else ""
            if 'SL NO' in val: col_map['SL No.'] = cell.column
            elif 'PRODUCT NAME' in val: col_map['Product Name'] = cell.column
            elif 'PRODUCT CODE' in val: col_map['Product Code'] = cell.column
            elif 'TYPE' in val: col_map['Type'] = cell.column
            elif 'DIAMETER' in val: col_map['Diameter'] = cell.column
            elif 'HEIGHT' in val: col_map['Height'] = cell.column
            elif 'LENGTH' in val: col_map['Length'] = cell.column
            elif 'WIDTH' in val: col_map['Width'] = cell.column
            elif 'POWER' in val: col_map['Power (W)'] = cell.column
            elif 'LUMENS' in val: col_map['Lumens (L/W)'] = cell.column
            elif 'MOUNTING' in val: col_map['Mounting'] = cell.column
            elif 'LIGHT DISTRIBUTION' in val: col_map['Light Distribution'] = cell.column
            elif 'LIGHT SOURCE' in val: col_map['Light Source'] = cell.column
            elif 'IP' in val: col_map['IP'] = cell.column
            elif 'CRI' in val: col_map['CRI'] = cell.column
            elif 'CCT' in val: col_map['CCT (K)'] = cell.column
            elif 'COMMENTS' in val: col_map['Comments'] = cell.column

        # State for propagating merged values
        curr_sl = ""
        curr_name = ""
        curr_type = ""
        curr_mounting = ""
        curr_dist = ""
        curr_source = ""
        curr_ip = ""
        curr_cri = ""
        curr_cct = ""
        curr_comments = ""
        curr_height = ""
        curr_diameter = ""
        curr_width = ""

        for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
            row_data = {}
            has_product_code = False
            
            # Extract basic info
            prod_code = sheet.cell(row=row_idx, column=col_map.get('Product Code', 0)).value if col_map.get('Product Code') else None
            
            # If no product code and no other data, skip (might be empty row)
            if prod_code is None:
                # Check if it's a completely empty row
                is_empty = True
                for info in col_map.values():
                    if sheet.cell(row=row_idx, column=info).value is not None:
                        is_empty = False
                        break
                if is_empty: continue
            
            # Update current state if new values available
            def get_val(col_name):
                col = col_map.get(col_name)
                if not col: return None
                val = sheet.cell(row=row_idx, column=col).value
                if isinstance(val, str):
                    return val.strip().replace('\n', ' ').replace('\r', ' ')
                return val

            temp_sl = get_val('SL No.')
            if temp_sl: curr_sl = temp_sl
            
            temp_name = get_val('Product Name')
            if temp_name: curr_name = temp_name
            
            temp_type = get_val('Type')
            if temp_type: curr_type = temp_type

            # Variations can also be merged
            temp_height = get_val('Height')
            if temp_height: curr_height = temp_height
            
            temp_diameter = get_val('Diameter')
            if temp_diameter: curr_diameter = temp_diameter
            
            temp_width = get_val('Width')
            if temp_width: curr_width = temp_width

            temp_mounting = get_val('Mounting')
            if temp_mounting: curr_mounting = temp_mounting
            
            temp_dist = get_val('Light Distribution')
            if temp_dist: curr_dist = temp_dist
            
            temp_source = get_val('Light Source')
            if temp_source: curr_source = temp_source
            
            temp_ip = get_val('IP')
            if temp_ip: curr_ip = temp_ip
            
            temp_cri = get_val('CRI')
            if temp_cri: curr_cri = temp_cri
            
            temp_cct = get_val('CCT (K)')
            if temp_cct: curr_cct = temp_cct
            
            temp_comments = get_val('Comments')
            if temp_comments: curr_comments = temp_comments

            # Only add if there's a product code OR a name (to avoid empty variation rows)
            if prod_code or curr_name:
                row_dict = {
                    'Category': sheet_name,
                    'SL No.': curr_sl,
                    'Product Name': curr_name,
                    'Product Code': prod_code if prod_code else "",
                    'Type': curr_type,
                    'Diameter': curr_diameter if not get_val('Diameter') else temp_diameter,
                    'Height': curr_height if not get_val('Height') else temp_height,
                    'Length': get_val('Length'),
                    'Width': curr_width if not get_val('Width') else temp_width,
                    'Power (W)': get_val('Power (W)'),
                    'Lumens (L/W)': get_val('Lumens (L/W)'),
                    'Mounting': curr_mounting,
                    'Light Distribution': curr_dist,
                    'Light Source': curr_source,
                    'IP': curr_ip,
                    'CRI': curr_cri,
                    'CCT (K)': curr_cct,
                    'Comments': curr_comments
                }
                # Fix for merged height/diameter: if the current row has a specific value, usage is clear.
                # If it's empty, use the 'curr' version which propagates from previous rows in that product block.
                
                # To be safer, only keep rows that have at least a Product Code or some technical spec
                if prod_code or row_dict['Power (W)'] or row_dict['Length']:
                    all_data.append(row_dict)

    # Write to CSV
    with open(output_csv, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(all_data)

if __name__ == "__main__":
    process_excel('d:/ConnexWebsite/CONNEX_PRODUCT_SHEET.xlsx', 'd:/ConnexWebsite/CONNEX_PRODUCTS.csv')
    print("Done! File saved to d:/ConnexWebsite/CONNEX_PRODUCTS.csv")
