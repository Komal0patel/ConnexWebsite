
import openpyxl
import json
import os
import re

def clean_val(val):
    if val is None: return None
    if isinstance(val, str):
        val = val.strip().replace('\n', ' ').replace('\r', ' ')
        if val == "": return None
    return val

def to_num(val):
    if val is None: return None
    if isinstance(val, (int, float)): return val
    # Try to extract number from string (e.g., "CRI 80" -> 80)
    match = re.search(r'(\d+)', str(val))
    if match:
        return int(match.group(1))
    return val

def process_excel_to_json(file_path, output_json):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    products_map = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"Processing sheet: {sheet_name}")
        
        header_row_idx = -1
        for row in sheet.iter_rows(min_row=1, max_row=10):
            row_values = [cell.value for cell in row]
            if any(val and isinstance(val, str) and 'PRODUCT CODE' in val.upper() for val in row_values):
                header_row_idx = row[0].row
                break
        
        if header_row_idx == -1: continue
            
        col_map = {}
        for cell in sheet[header_row_idx]:
            val = str(cell.value).strip().upper() if cell.value else ""
            if 'SL NO' in val: col_map['sl_no'] = cell.column
            elif 'PRODUCT NAME' in val: col_map['name'] = cell.column
            elif 'PRODUCT CODE' in val: col_map['code'] = cell.column
            elif 'TYPE' in val: col_map['type'] = cell.column
            elif 'DIAMETER' in val: col_map['diameter'] = cell.column
            elif 'HEIGHT' in val: col_map['height'] = cell.column
            elif 'LENGTH' in val: col_map['length'] = cell.column
            elif 'WIDTH' in val: col_map['width'] = cell.column
            elif 'POWER' in val: col_map['power'] = cell.column
            elif 'LUMENS' in val: col_map['lumens'] = cell.column
            elif 'MOUNTING' in val: col_map['mounting'] = cell.column
            elif 'LIGHT DISTRIBUTION' in val: col_map['light_dist'] = cell.column
            elif 'LIGHT SOURCE' in val: col_map['light_source'] = cell.column
            elif 'IP' in val: col_map['ip'] = cell.column
            elif 'CRI' in val: col_map['cri'] = cell.column
            elif 'CCT' in val: col_map['cct'] = cell.column
            elif 'COMMENTS' in val: col_map['comments'] = cell.column

        curr_name = ""
        curr_category = sheet_name
        
        # Temp state for merged specs
        state = {
            'mounting': "", 'light_dist': "", 'light_source': "",
            'ip': None, 'cri': None, 'comments': ""
        }

        for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
            def get_raw(key):
                col = col_map.get(key)
                return clean_val(sheet.cell(row=row_idx, column=col).value) if col else None

            name = get_raw('name')
            if name: curr_name = name
            
            if not curr_name: continue

            # Update shared state
            for key in state.keys():
                val = get_raw(key)
                if val: state[key] = val

            # Product key (Name + Category)
            prod_key = f"{curr_name}_{curr_category}"
            
            if prod_key not in products_map:
                products_map[prod_key] = {
                    "product_name": curr_name,
                    "category": curr_category,
                    "image": f"{curr_name.lower().replace(' ', '_')}.png",
                    "mounting": state['mounting'],
                    "light_distribution": state['light_dist'],
                    "light_source": state['light_source'],
                    "ip": to_num(state['ip']),
                    "cri": to_num(state['cri']),
                    "comments": state['comments'],
                    "variants": []
                }

            # Create variant
            code = get_raw('code')
            power = get_raw('power')
            
            # Only add if there is a code or some tech spec
            if code or power:
                variant = {
                    "code": code,
                    "type": get_raw('type'),
                    "diameter": to_num(get_raw('diameter')),
                    "height": to_num(get_raw('height')),
                    "length": to_num(get_raw('length')),
                    "width": to_num(get_raw('width')),
                    "power": to_num(power),
                    "lumens": to_num(get_raw('lumens')),
                    "cct": to_num(get_raw('cct'))
                }
                # Remove nulls from variant for cleaner data
                variant = {k: v for k, v in variant.items() if v is not None}
                products_map[prod_key]["variants"].append(variant)

    # Convert map to list
    final_data = list(products_map.values())
    
    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump(final_data, f, indent=2)

if __name__ == "__main__":
    process_excel_to_json('d:/ConnexWebsite/CONNEX_PRODUCT_SHEET.xlsx', 'd:/ConnexWebsite/CONNEX_PRODUCTS.json')
    print("Done! File saved to d:/ConnexWebsite/CONNEX_PRODUCTS.json")
